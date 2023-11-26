import datetime
import os.path
import random
from typing import Dict, List
import openpyxl

def generate_error_data(num_errors: int) -> List[Dict[str, str]]:
    """
    Для тестирования, генерирует случайные данные об ошибках схожие с теми, что приходят от станка.

    Attributes:
        num_errors (int): Количество ошибок для генерации.

    Returns:
        list: Список словарей с данными об ошибках, включая код ошибки и описание.
    """
    errors = []
    for _ in range(num_errors):
        error_code = f'{random.randint(0, 99):04d}'
        error_description = f'Error description for {error_code}'

        errors.append({
            'error_code': error_code,
            'error_description': error_description,
        })

    return errors

def save_to_excel(errors: List[Dict[str, str]], filename=f"error_log_"):
    """
    Сохраняет данные об ошибках в файл Excel.

    Attributes:
        errors (List[Dict[str, str]]): Список словарей с данными об ошибках.
        filename (str): Имя файла для сохранения данных. Опционально, по умолчанию - error_log_
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['Error Code', 'Error Description']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    for row_num, error in enumerate(errors, 2):
        ws.cell(row=row_num, column=1, value=error['error_code'])
        ws.cell(row=row_num, column=2, value=error['error_description'])
    wb.save(filename+f"{datetime.datetime.now().strftime('%d.%m.%Y_%H_%M')}.xlsx")

error_data = generate_error_data(100)
file = os.path.join('data', 'error_log')
save_to_excel(error_data, filename=file)
